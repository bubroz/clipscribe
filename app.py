import streamlit as st
import asyncio
from pathlib import Path
import os
import time
import pandas as pd
import json
from typing import Dict, List, Any, Optional

from src.clipscribe.retrievers import VideoIntelligenceRetriever, UniversalVideoClient
from src.clipscribe.utils.performance import PerformanceMonitor
from src.clipscribe.utils.performance_dashboard import create_performance_dashboard
from src.clipscribe.extractors.model_manager import model_manager
import logging

logger = logging.getLogger(__name__)

# --- Page Config ---
st.set_page_config(
    page_title="ClipScribe Web UI",
    page_icon="🚀",
    layout="wide",
)

# --- App State ---
def initialize_session_state():
    """Initialize all session state variables."""
    defaults = {
        'processing': False,
        'video_result': None,
        'output_dir': None,
        'research_results': [],
        'research_in_progress': False,
        'processed_results': {},
        'research_df_data': [],
        'batch_progress': {},
        'performance_monitor': None,
        'entity_analysis_data': None,
        'model_cache_stats': {},
        'batch_start_time': None,
        'total_batch_time': None
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

async def process_video_url(url, output_dir, mode, use_cache, clean_graph):
    """The main async function to process a video."""
    st.session_state.processing = True
    st.session_state.video_result = None
    
    status_placeholder = st.empty()
    progress_bar = st.progress(0, "Initializing...")

    def streamlit_progress_hook(update):
        description = update.get("description", "Processing...")
        progress_value = update.get("progress", 0)
        status_placeholder.info(f"⚙️ {description}")
        progress_bar.progress(progress_value, text=description)

    try:
        # Initialize performance monitoring
        if not st.session_state.performance_monitor:
            st.session_state.performance_monitor = PerformanceMonitor(output_dir)
        
        perf_event = st.session_state.performance_monitor.start_timer("single_video_processing", url=url, mode=mode)
        
        retriever = VideoIntelligenceRetriever(
            use_cache=use_cache,
            mode=mode,
            output_dir=output_dir,
            progress_hook=streamlit_progress_hook
        )
        retriever.clean_graph = clean_graph
        
        result = await retriever.process_url(url)
        
        if result:
            streamlit_progress_hook({"description": "Saving output files...", "progress": 95})
            retriever.save_all_formats(
                result, 
                output_dir=str(output_dir),
                include_chimera_format=True
            )
            
            # Record performance metrics
            st.session_state.performance_monitor.record_metric(
                "entities_extracted", 
                len(result.entities),
                url=url
            )
            st.session_state.performance_monitor.record_metric(
                "relationships_extracted", 
                len(getattr(result, 'relationships', [])),
                url=url
            )
            
        st.session_state.video_result = result
        st.session_state.output_dir = output_dir
        
        st.session_state.performance_monitor.stop_timer(perf_event)

    except Exception as e:
        st.error(f"An error occurred: {e}")
        if 'perf_event' in locals():
            st.session_state.performance_monitor.stop_timer(perf_event)
    finally:
        st.session_state.processing = False
        progress_bar.empty()
        status_placeholder.empty()

async def run_research_search(query: str, search_type: str, max_results: int, period: str, sort_by: str):
    """Runs the backend search for videos."""
    st.session_state.research_in_progress = True
    st.session_state.research_results = []
    st.session_state.research_df_data = [] # Clear previous results
    
    try:
        video_client = UniversalVideoClient()
        search_results = []

        is_channel_search = search_type == "Channel URL"

        if is_channel_search:
            if "youtube.com/" not in query and "youtu.be/" not in query:
                st.error("Please enter a valid YouTube channel URL.")
                return
            search_results = await video_client.search_channel(query, max_results=max_results, sort_by=sort_by)
        else:
            search_results = await video_client.search_videos(query, max_results=max_results, period=period)
        
        st.session_state.research_results = search_results
    except Exception as e:
        st.error(f"An error occurred during search: {e}")
    finally:
        st.session_state.research_in_progress = False

async def process_video_for_batch(video_meta, sidebar_config, progress_callback=None):
    """Processes a single video for a batch job with enhanced progress tracking."""
    video_output_dir = Path(f"output/research/youtube_{video_meta.video_id}")
    video_output_dir.mkdir(parents=True, exist_ok=True)

    # Initialize performance monitoring if not exists
    if not st.session_state.performance_monitor:
        st.session_state.performance_monitor = PerformanceMonitor(Path("output/research"))
    
    perf_event = st.session_state.performance_monitor.start_timer(
        "batch_video_processing", 
        url=video_meta.url, 
        video_id=video_meta.video_id,
        mode=sidebar_config['processing_mode']
    )

    retriever = VideoIntelligenceRetriever(
        use_cache=sidebar_config['use_cache'],
        mode=sidebar_config['processing_mode'],
        output_dir=video_output_dir,
        progress_hook=progress_callback
    )
    retriever.clean_graph = sidebar_config['clean_graph']
    
    try:
        result = await retriever.process_url(video_meta.url)
        if result:
            retriever.save_all_formats(
                result, 
                output_dir=str(video_output_dir),
                include_chimera_format=True
            )
            st.session_state.processed_results[video_meta.url] = result
            
            # Record detailed metrics
            st.session_state.performance_monitor.record_metric(
                "batch_entities_extracted", 
                len(result.entities),
                url=video_meta.url,
                video_id=video_meta.video_id
            )
            st.session_state.performance_monitor.record_metric(
                "batch_relationships_extracted", 
                len(getattr(result, 'relationships', [])),
                url=video_meta.url,
                video_id=video_meta.video_id
            )
            
        st.session_state.performance_monitor.stop_timer(perf_event)
        return video_meta.url, "Complete", result
    except Exception as e:
        logger.error(f"Error processing {video_meta.url}: {e}")
        st.session_state.performance_monitor.stop_timer(perf_event)
        return video_meta.url, "Error", None

async def run_batch_processing(sidebar_config):
    """Runs the full batch processing job with enhanced progress tracking and analytics."""
    st.session_state.research_in_progress = True
    st.session_state.batch_start_time = time.time()
    
    results_to_process = [
        video for video in st.session_state.research_results 
        if video.url not in st.session_state.processed_results
    ]
    
    if not results_to_process:
        st.toast("All videos have already been processed.")
        st.session_state.research_in_progress = False
        return

    # Initialize batch progress tracking
    st.session_state.batch_progress = {
        'total': len(results_to_process),
        'completed': 0,
        'errors': 0,
        'current_video': '',
        'start_time': time.time()
    }

    # Update the UI to show "Processing" for all pending videos
    for row in st.session_state.research_df_data:
        if row["Status"] == "Pending":
            row["Status"] = "Processing..."
    
    # Create progress display
    progress_container = st.container()
    with progress_container:
        progress_text = f"Processing {len(results_to_process)} videos..."
        main_progress = st.progress(0, text=progress_text)
        status_text = st.empty()
        
        # Model cache stats display
        cache_stats = model_manager.get_cache_info()
        st.session_state.model_cache_stats = cache_stats
        
        if cache_stats['model_count'] > 0:
            st.info(f"🚀 Using cached models: {', '.join(cache_stats['cached_models'])}")
    
    # Process videos with detailed progress tracking
    for i, video_meta in enumerate(results_to_process):
        st.session_state.batch_progress['current_video'] = video_meta.title
        
        def progress_callback(update):
            description = update.get("description", "Processing...")
            status_text.text(f"📹 {video_meta.title[:50]}... - {description}")
        
        url, status, result = await process_video_for_batch(video_meta, sidebar_config, progress_callback)
        
        # Update progress tracking
        if status == "Complete":
            st.session_state.batch_progress['completed'] += 1
        else:
            st.session_state.batch_progress['errors'] += 1
        
        # Update UI
        for row in st.session_state.research_df_data:
            if row["URL"] == url:
                row["Status"] = status
                if result:
                    row["Entities"] = len(result.entities)
                    row["Relationships"] = len(getattr(result, 'relationships', []))
        
        progress_value = (i + 1) / len(results_to_process)
        elapsed_time = time.time() - st.session_state.batch_progress['start_time']
        eta = (elapsed_time / (i + 1)) * (len(results_to_process) - i - 1) if i > 0 else 0
        
        main_progress.progress(
            progress_value, 
            text=f"Processed {i + 1} of {len(results_to_process)} videos (ETA: {eta:.0f}s)"
        )
        
        # Rerun to update the data editor
        st.rerun()

    # Final batch statistics
    st.session_state.total_batch_time = time.time() - st.session_state.batch_start_time
    st.session_state.research_in_progress = False
    
    # Save performance report
    if st.session_state.performance_monitor:
        st.session_state.performance_monitor.save_report()
    
    st.rerun()

def analyze_entity_sources(processed_results: Dict[str, Any]) -> Dict[str, Any]:
    """Analyze entity sources across all processed videos."""
    if not processed_results:
        return {}
    
    analysis = {
        'total_videos': len(processed_results),
        'total_entities': 0,
        'source_distribution': {},
        'type_distribution': {},
        'confidence_stats': {},
        'videos_analysis': []
    }
    
    all_confidences = []
    
    for url, result in processed_results.items():
        video_analysis = {
            'url': url,
            'title': result.metadata.title if hasattr(result, 'metadata') else 'Unknown',
            'entity_count': len(result.entities),
            'source_breakdown': {},
            'type_breakdown': {}
        }
        
        for entity in result.entities:
            analysis['total_entities'] += 1
            all_confidences.append(entity.confidence)
            
            # Source analysis
            source = "Unknown"
            if hasattr(entity, 'properties') and entity.properties:
                source = entity.properties.get('source', 'Unknown')
            
            analysis['source_distribution'][source] = analysis['source_distribution'].get(source, 0) + 1
            video_analysis['source_breakdown'][source] = video_analysis['source_breakdown'].get(source, 0) + 1
            
            # Type analysis
            entity_type = entity.type
            analysis['type_distribution'][entity_type] = analysis['type_distribution'].get(entity_type, 0) + 1
            video_analysis['type_breakdown'][entity_type] = video_analysis['type_breakdown'].get(entity_type, 0) + 1
        
        analysis['videos_analysis'].append(video_analysis)
    
    # Confidence statistics
    if all_confidences:
        analysis['confidence_stats'] = {
            'average': sum(all_confidences) / len(all_confidences),
            'min': min(all_confidences),
            'max': max(all_confidences),
            'high_confidence_count': len([c for c in all_confidences if c > 0.8])
        }
    
    return analysis

def generate_excel_export(analysis_data: Dict[str, Any]) -> bytes:
    """Generate Excel export of entity analysis data."""
    import io
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Create workbook
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Videos", analysis_data['total_videos']],
        ["Total Entities", analysis_data['total_entities']],
        ["Average Confidence", f"{analysis_data.get('confidence_stats', {}).get('average', 0):.3f}"],
        ["High Confidence Count", analysis_data.get('confidence_stats', {}).get('high_confidence_count', 0)]
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Source distribution sheet
    if analysis_data['source_distribution']:
        ws_sources = wb.create_sheet("Source Distribution")
        source_df = pd.DataFrame([
            {"Extraction Method": source, "Entity Count": count, 
             "Percentage": f"{(count/analysis_data['total_entities']*100):.1f}%"}
            for source, count in sorted(analysis_data['source_distribution'].items(), 
                                      key=lambda x: x[1], reverse=True)
        ])
        
        for r in dataframe_to_rows(source_df, index=False, header=True):
            ws_sources.append(r)
    
    # Type distribution sheet
    if analysis_data['type_distribution']:
        ws_types = wb.create_sheet("Entity Types")
        type_df = pd.DataFrame([
            {"Entity Type": entity_type, "Count": count}
            for entity_type, count in sorted(analysis_data['type_distribution'].items(), 
                                           key=lambda x: x[1], reverse=True)
        ])
        
        for r in dataframe_to_rows(type_df, index=False, header=True):
            ws_types.append(r)
    
    # Per-video analysis sheet
    if analysis_data.get('videos_analysis'):
        ws_videos = wb.create_sheet("Per-Video Analysis")
        video_data = []
        for video in analysis_data['videos_analysis']:
            video_data.append({
                "Title": video['title'],
                "Entity Count": video['entity_count'],
                "Top Source": max(video['source_breakdown'].items(), key=lambda x: x[1])[0] if video['source_breakdown'] else "None",
                "Top Type": max(video['type_breakdown'].items(), key=lambda x: x[1])[0] if video['type_breakdown'] else "None"
            })
        
        video_df = pd.DataFrame(video_data)
        for r in dataframe_to_rows(video_df, index=False, header=True):
            ws_videos.append(r)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_csv_export(analysis_data: Dict[str, Any]) -> str:
    """Generate CSV export of entity analysis data."""
    csv_data = []
    
    if analysis_data.get('videos_analysis'):
        # Per-video CSV
        for video in analysis_data['videos_analysis']:
            row = {
                'Video Title': video['title'],
                'Entity Count': video['entity_count'],
            }
            
            # Add source breakdown
            for source, count in video['source_breakdown'].items():
                row[f'{source}_Count'] = count
            
            # Add type breakdown (top 5 only to keep CSV manageable)
            top_types = sorted(video['type_breakdown'].items(), key=lambda x: x[1], reverse=True)[:5]
            for i, (entity_type, count) in enumerate(top_types):
                row[f'Top_Type_{i+1}'] = f"{entity_type} ({count})"
            
            csv_data.append(row)
    else:
        # Single video or summary CSV
        csv_data = [{
            'Total Videos': analysis_data['total_videos'],
            'Total Entities': analysis_data['total_entities'],
            'Average Confidence': analysis_data.get('confidence_stats', {}).get('average', 0),
            'High Confidence Count': analysis_data.get('confidence_stats', {}).get('high_confidence_count', 0)
        }]
    
    df = pd.DataFrame(csv_data)
    return df.to_csv(index=False)

def generate_markdown_export(analysis_data: Dict[str, Any]) -> str:
    """Generate Markdown export of entity analysis data."""
    from datetime import datetime
    
    report = ["# Entity Source Analysis Report\n"]
    report.append(f"**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # Summary
    report.append("## 📊 Summary\n")
    report.append(f"- **Total Videos Analyzed**: {analysis_data['total_videos']}")
    report.append(f"- **Total Entities Extracted**: {analysis_data['total_entities']}")
    
    if analysis_data.get('confidence_stats'):
        conf_stats = analysis_data['confidence_stats']
        report.append(f"- **Average Confidence**: {conf_stats.get('average', 0):.3f}")
        report.append(f"- **High Confidence Entities**: {conf_stats.get('high_confidence_count', 0)}")
    report.append("")
    
    # Source distribution
    if analysis_data['source_distribution']:
        report.append("## 🔍 Extraction Method Performance\n")
        report.append("| Method | Entity Count | Percentage |\n")
        report.append("|--------|--------------|------------|\n")
        
        total_entities = analysis_data['total_entities']
        for source, count in sorted(analysis_data['source_distribution'].items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total_entities) * 100 if total_entities > 0 else 0
            report.append(f"| {source} | {count} | {percentage:.1f}% |\n")
        report.append("")
    
    # Entity type distribution
    if analysis_data['type_distribution']:
        report.append("## 🏷️ Top Entity Types\n")
        report.append("| Entity Type | Count |\n")
        report.append("|-------------|-------|\n")
        
        top_types = sorted(analysis_data['type_distribution'].items(), key=lambda x: x[1], reverse=True)[:10]
        for entity_type, count in top_types:
            report.append(f"| {entity_type} | {count} |\n")
        report.append("")
    
    # Per-video analysis
    if analysis_data.get('videos_analysis'):
        report.append("## 📹 Per-Video Analysis\n")
        report.append("| Video Title | Entities | Top Source | Top Type |\n")
        report.append("|-------------|----------|------------|----------|\n")
        
        for video in analysis_data['videos_analysis']:
            title = video['title'][:50] + "..." if len(video['title']) > 50 else video['title']
            entities = video['entity_count']
            top_source = max(video['source_breakdown'].items(), key=lambda x: x[1])[0] if video['source_breakdown'] else "None"
            top_type = max(video['type_breakdown'].items(), key=lambda x: x[1])[0] if video['type_breakdown'] else "None"
            
            report.append(f"| {title} | {entities} | {top_source} | {top_type} |\n")
        report.append("")
    
    return "\n".join(report)

# --- UI Rendering Functions ---
async def render_single_video_tab(sidebar_config):
    """Renders the UI for the single video analysis tab."""
    st.markdown("Transform any video into structured, searchable knowledge.")
    url = st.text_input("Enter Video URL", placeholder="https://www.youtube.com/watch?v=...")

    if st.button("Analyze Video", type="primary", disabled=st.session_state.get('processing', False)):
        if url:
            timestamp = time.strftime("%Y%m%d-%H%M%S")
            video_id = url.split("=")[-1] if "youtube" in url else url.split("/")[-1]
            output_dir = Path(f"output/streamlit_{timestamp}_{video_id}")
            output_dir.mkdir(parents=True, exist_ok=True)
            
            await process_video_url(
                url, output_dir,
                mode=sidebar_config['processing_mode'],
                use_cache=sidebar_config['use_cache'],
                clean_graph=sidebar_config['clean_graph']
            )
        else:
            st.warning("Please enter a URL.")

    if st.session_state.get('video_result'):
        render_video_results(st.session_state.video_result, st.session_state.output_dir)

async def render_research_tab(sidebar_config):
    """Renders the UI for the research tab with enhanced batch processing."""
    st.header("🔬 Research")
    st.markdown("Perform batch analysis by searching for videos by topic or channel.")

    search_type = st.radio("Search Type", ["Topic", "Channel URL"], horizontal=True)
    query = st.text_input("Search Query", placeholder="e.g., 'AI in healthcare' or 'https://www.youtube.com/@pbsnewshour'")

    col1, col2, col3 = st.columns(3)
    with col1:
        max_results = st.number_input("Max Videos", min_value=1, max_value=50, value=5)
    with col2:
        period = st.selectbox("Time Period", ["any", "hour", "today", "week", "month", "year"], index=0, help="Filter by upload date (topic search only).")
    with col3:
        sort_by = st.selectbox("Sort By", ["relevance", "view_count", "upload_date", "rating"], index=0, help="Sort order (channel search only).")
    
    if st.button("Search for Videos", type="primary", disabled=st.session_state.research_in_progress):
        if query:
            with st.spinner(f"Searching for {search_type}: {query}..."):
                await run_research_search(query, search_type, max_results, period, sort_by)
            st.rerun()
        else:
            st.warning("Please enter a search query.")
    
    if st.session_state.research_results or st.session_state.research_df_data:
        await render_research_results(sidebar_config)

async def render_research_results(sidebar_config):
    """Renders the research results with enhanced batch processing and analytics."""
    if not st.session_state.research_df_data and st.session_state.research_results:
        st.session_state.research_df_data = [
            {
                "Title": r.title, 
                "Channel": r.channel, 
                "Duration (s)": r.duration,
                "Views": r.view_count, 
                "URL": r.url, 
                "Status": "Pending",
                "Entities": 0,
                "Relationships": 0
            } for r in st.session_state.research_results
        ]
    
    if not st.session_state.research_df_data:
        st.info("No videos found for your query.")
        return

    st.success(f"Found {len(st.session_state.research_df_data)} videos.")
    
    # Enhanced data display with metrics
    df = pd.DataFrame(st.session_state.research_df_data)
    
    # Add summary metrics if batch processing has been run
    if st.session_state.processed_results:
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            completed = len([r for r in st.session_state.research_df_data if r["Status"] == "Complete"])
            st.metric("Completed", completed, delta=completed - len(st.session_state.processed_results))
        
        with col2:
            total_entities = sum(r.get("Entities", 0) for r in st.session_state.research_df_data)
            st.metric("Total Entities", total_entities)
        
        with col3:
            total_relationships = sum(r.get("Relationships", 0) for r in st.session_state.research_df_data)
            st.metric("Total Relationships", total_relationships)
        
        with col4:
            if st.session_state.total_batch_time:
                st.metric("Batch Time", f"{st.session_state.total_batch_time:.1f}s")
    
    st.data_editor(df, key="research_results_editor", use_container_width=True,
                   disabled=["Title", "Channel", "Duration (s)", "Views", "URL", "Status", "Entities", "Relationships"], 
                   hide_index=True)

    # Processing controls
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Process All Videos", type="primary", disabled=st.session_state.research_in_progress):
            await run_batch_processing(sidebar_config)
    
    with col2:
        if st.session_state.processed_results and st.button("Analyze Results", type="secondary"):
            st.session_state.entity_analysis_data = analyze_entity_sources(st.session_state.processed_results)
            st.rerun()
    
    # Display batch progress if processing
    if st.session_state.research_in_progress and st.session_state.batch_progress:
        progress_info = st.session_state.batch_progress
        st.info(f"Processing: {progress_info['completed']}/{progress_info['total']} complete, {progress_info['errors']} errors")
    
    # Display entity analysis results
    if st.session_state.entity_analysis_data:
        render_entity_analysis(st.session_state.entity_analysis_data)

def render_entity_analysis(analysis_data: Dict[str, Any]):
    """Render comprehensive entity source analysis with v2.12.0 enhancements."""
    st.header("📊 Entity Source Analysis")
    
    # Import Plotly for advanced visualizations
    try:
        import plotly.graph_objects as go
        import plotly.express as px
        from plotly.subplots import make_subplots
        PLOTLY_AVAILABLE = True
    except ImportError:
        PLOTLY_AVAILABLE = False
        st.warning("⚠️ Plotly not available. Install with: pip install plotly")
    
    # Overview metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Videos Analyzed", analysis_data['total_videos'])
    with col2:
        st.metric("Total Entities Extracted", analysis_data['total_entities'])
    with col3:
        if analysis_data.get('confidence_stats'):
            avg_conf = analysis_data['confidence_stats']['average']
            st.metric("Average Confidence", f"{avg_conf:.2f}")
    with col4:
        high_conf_count = analysis_data.get('confidence_stats', {}).get('high_confidence_count', 0)
        high_conf_ratio = high_conf_count / analysis_data['total_entities'] if analysis_data['total_entities'] > 0 else 0
        st.metric("High Confidence", f"{high_conf_ratio:.1%}")
    
    # Export controls
    st.subheader("📥 Export Analysis Results")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📊 Download Excel Report"):
            excel_data = generate_excel_export(analysis_data)
            st.download_button(
                label="📥 Download Excel File",
                data=excel_data,
                file_name=f"entity_analysis_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col2:
        if st.button("📄 Download CSV Report"):
            csv_data = generate_csv_export(analysis_data)
            st.download_button(
                label="📥 Download CSV File",
                data=csv_data,
                file_name=f"entity_analysis_{time.strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
    
    with col3:
        if st.button("📋 Download Markdown Report"):
            md_data = generate_markdown_export(analysis_data)
            st.download_button(
                label="📥 Download Markdown File",
                data=md_data,
                file_name=f"entity_analysis_{time.strftime('%Y%m%d_%H%M%S')}.md",
                mime="text/markdown"
            )
    
    # Source distribution with enhanced visualizations
    st.subheader("🔍 Extraction Method Performance")
    
    if analysis_data['source_distribution']:
        source_df = pd.DataFrame([
            {"Extraction Method": source, "Entity Count": count, 
             "Percentage": f"{(count/analysis_data['total_entities']*100):.1f}%"}
            for source, count in sorted(analysis_data['source_distribution'].items(), 
                                      key=lambda x: x[1], reverse=True)
        ])
        
        # Enhanced visualizations with Plotly
        if PLOTLY_AVAILABLE:
            col1, col2 = st.columns(2)
            
            with col1:
                # Interactive pie chart
                fig_pie = go.Figure(data=[go.Pie(
                    labels=list(analysis_data['source_distribution'].keys()),
                    values=list(analysis_data['source_distribution'].values()),
                    hole=0.3,
                    textinfo='label+percent',
                    marker_colors=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7']
                )])
                
                fig_pie.update_layout(
                    title="Extraction Method Distribution",
                    annotations=[dict(text='Methods', x=0.5, y=0.5, font_size=16, showarrow=False)]
                )
                
                st.plotly_chart(fig_pie, use_container_width=True)
            
            with col2:
                # Interactive bar chart
                fig_bar = go.Figure(data=[
                    go.Bar(
                        x=list(analysis_data['source_distribution'].keys()),
                        y=list(analysis_data['source_distribution'].values()),
                        marker_color='lightblue',
                        text=list(analysis_data['source_distribution'].values()),
                        textposition='auto'
                    )
                ])
                
                fig_bar.update_layout(
                    title="Entity Count by Method",
                    xaxis_title="Extraction Method",
                    yaxis_title="Entity Count"
                )
                
                st.plotly_chart(fig_bar, use_container_width=True)
        else:
            # Fallback to simple charts
            col1, col2 = st.columns([2, 1])
            with col1:
                st.dataframe(source_df, use_container_width=True)
            with col2:
                st.bar_chart(source_df.set_index("Extraction Method")["Entity Count"])
    
    # Entity type distribution with visualization
    st.subheader("🏷️ Entity Type Distribution")
    if analysis_data['type_distribution']:
        type_df = pd.DataFrame([
            {"Entity Type": entity_type, "Count": count}
            for entity_type, count in sorted(analysis_data['type_distribution'].items(), 
                                           key=lambda x: x[1], reverse=True)[:10]  # Top 10
        ])
        
        if PLOTLY_AVAILABLE:
            # Interactive horizontal bar chart for entity types
            fig_types = go.Figure(data=[
                go.Bar(
                    x=type_df["Count"],
                    y=type_df["Entity Type"],
                    orientation='h',
                    marker_color='lightcoral',
                    text=type_df["Count"],
                    textposition='auto'
                )
            ])
            
            fig_types.update_layout(
                title="Top 10 Entity Types",
                xaxis_title="Count",
                yaxis_title="Entity Type",
                height=400
            )
            
            st.plotly_chart(fig_types, use_container_width=True)
        else:
            st.dataframe(type_df, use_container_width=True)
    
    # Per-video breakdown
    st.subheader("📹 Per-Video Analysis")
    video_summary = []
    for video in analysis_data['videos_analysis']:
        video_summary.append({
            "Title": video['title'][:50] + "..." if len(video['title']) > 50 else video['title'],
            "Entities": video['entity_count'],
            "Top Source": max(video['source_breakdown'].items(), key=lambda x: x[1])[0] if video['source_breakdown'] else "None",
            "Top Type": max(video['type_breakdown'].items(), key=lambda x: x[1])[0] if video['type_breakdown'] else "None"
        })
    
    if video_summary:
        video_df = pd.DataFrame(video_summary)
        st.dataframe(video_df, use_container_width=True)
    
    # Model performance insights
    if st.session_state.model_cache_stats:
        st.subheader("⚡ Model Performance")
        cache_stats = st.session_state.model_cache_stats
        
        st.info(f"**Model Caching Active**: {cache_stats['model_count']} models cached")
        if cache_stats['cached_models']:
            for model in cache_stats['cached_models']:
                st.text(f"• {model}")
        
        if st.session_state.total_batch_time and analysis_data['total_videos'] > 1:
            avg_time_per_video = st.session_state.total_batch_time / analysis_data['total_videos']
            st.metric("Average Time per Video", f"{avg_time_per_video:.1f}s")

def render_video_results(result, output_dir):
    """Renders the results of a single video analysis with entity source information."""
    st.success("Analysis Complete!")
    st.header(result.metadata.title)
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Duration", f"{int(result.metadata.duration // 60)}m {int(result.metadata.duration % 60)}s")
    col2.metric("Entities", len(result.entities))
    col3.metric("Relationships", len(getattr(result, 'relationships', [])))

    with st.expander("📝 Executive Summary", expanded=True):
        st.markdown(result.summary)
    
    # Entity source breakdown for single video
    with st.expander("🔍 Entity Source Analysis"):
        if result.entities:
            entity_sources = {}
            for entity in result.entities:
                source = "Unknown"
                if hasattr(entity, 'properties') and entity.properties:
                    source = entity.properties.get('source', 'Unknown')
                entity_sources[source] = entity_sources.get(source, 0) + 1
            
            source_df = pd.DataFrame([
                {"Source": source, "Count": count}
                for source, count in entity_sources.items()
            ])
            st.dataframe(source_df, use_container_width=True)
        else:
            st.info("No entities found in this video.")

    with st.expander("📊 Markdown Report"):
        try:
            subdirectories = [d for d in output_dir.iterdir() if d.is_dir()]
            if subdirectories:
                report_path = subdirectories[0] / "report.md"
                with open(report_path, "r", encoding="utf-8") as f:
                    report_content = f.read()
                st.markdown(report_content, unsafe_allow_html=True)
            else:
                st.warning("Markdown report not found.")
        except FileNotFoundError:
            st.warning("Markdown report not found.")

    with st.expander("🗂️ Download Files"):
        subdirectories = [d for d in output_dir.iterdir() if d.is_dir()]
        if subdirectories:
            files_dir = subdirectories[0]
            for filename in os.listdir(files_dir):
                file_path = files_dir / filename
                if file_path.is_file():
                    with open(file_path, "rb") as f:
                        st.download_button(
                            label=f"Download {filename}", data=f, file_name=filename,
                            mime="application/octet-stream"
                        )
        else:
            st.warning("Could not find output directory with files.") 

async def main():
    """The main Streamlit application logic."""
    initialize_session_state()
    st.title("🚀 ClipScribe Web UI")
    st.caption("v2.12.0 with Advanced Visualizations, Excel Export & Performance Dashboards")

    # --- Sidebar Configuration ---
    st.sidebar.title("⚙️ Configuration")
    processing_mode = st.sidebar.selectbox(
        "Processing Mode", options=["audio", "video", "auto"], index=2,
        help="**Audio**: Fastest. **Video**: Slower. **Auto**: Detects best mode."
    )
    use_cache = st.sidebar.checkbox("Use Cache", value=True, help="Use cached results.")
    clean_graph = st.sidebar.checkbox("Clean Knowledge Graph", value=True, help="Refine graph with an AI pass.")
    
    # Model cache information in sidebar
    st.sidebar.subheader("🚀 Model Cache Status")
    cache_info = model_manager.get_cache_info()
    if cache_info['model_count'] > 0:
        st.sidebar.success(f"✅ {cache_info['model_count']} models cached")
        for model in cache_info['cached_models']:
            st.sidebar.text(f"• {model}")
    else:
        st.sidebar.info("No models cached yet")
    
    if st.sidebar.button("Clear Model Cache"):
        model_manager.clear_cache()
        st.sidebar.success("Model cache cleared!")
        st.rerun()
    
    sidebar_config = {
        "processing_mode": processing_mode,
        "use_cache": use_cache,
        "clean_graph": clean_graph
    }

    tab1, tab2, tab3 = st.tabs([
        "🔎 Single Video Analysis", 
        "🔬 Research & Batch Processing",
        "📊 Performance Dashboard"
    ])

    with tab1:
        await render_single_video_tab(sidebar_config)

    with tab2:
        await render_research_tab(sidebar_config)
    
    with tab3:
        render_performance_dashboard_tab()

def render_performance_dashboard_tab():
    """Render the dedicated performance dashboard tab."""
    from src.clipscribe.utils.performance_dashboard import create_performance_dashboard
    
    # Create performance dashboard instance
    output_dir = Path("output")
    dashboard = create_performance_dashboard(output_dir)
    
    # Render the full dashboard
    dashboard.render_dashboard()

if __name__ == "__main__":
    asyncio.run(main())